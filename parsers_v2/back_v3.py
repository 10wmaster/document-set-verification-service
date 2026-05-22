import io
import re
import enum
import os
from datetime import datetime, timedelta
from urllib.parse import quote  # Для безопасного кодирования имени файла
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

# Импорты для генерации стилизованного Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = FastAPI(
    title="OHS Expert Compliance Service API",
    description="MVP сервиса экспресс-аудита ЛНА по охране труда с учетом требований Приказа 772н и Постановления 2464"
)


# === 1. ПЕРЕЧИСЛЕНИЯ ===
class SeverityLevel(str, enum.Enum):
    CRITICAL = "CRITICAL"
    WARNING = "WARNING"
    RECOMMENDATION = "RECOMMEND"


# === 2. СХЕМЫ ДАННЫХ (Pydantic-модели) ===
class ValidationErrorItem(BaseModel):
    category: str
    severity: SeverityLevel
    message: str
    location: Optional[str] = None
    fine_equivalent: Optional[str] = "0 руб."
    legal_tip: Optional[str] = None


class DocumentChunk(BaseModel):
    text: str
    font: str
    size: float
    page: Optional[int] = None


class ValidationRequest(BaseModel):
    doc_type: str  # "instruction", "journal"
    document_data: List[DocumentChunk]


# === 3. БИЗНЕС-ЛОГИКА ===

def validate_gost_7_0_97(chunks: List[DocumentChunk]) -> List[ValidationErrorItem]:
    errors = []
    for idx, chunk in enumerate(chunks):
        if len(chunk.text.strip()) < 3:
            continue

        font_lower = chunk.font.lower()
        if "times" not in font_lower and "arial" not in font_lower:
            errors.append(ValidationErrorItem(
                category="Нормоконтроль (ГОСТ 7.0.97)",
                severity=SeverityLevel.WARNING,
                message=f"Нестандартный шрифт '{chunk.font}'.",
                location=f"Стр. {chunk.page if chunk.page else idx + 1}",
                fine_equivalent="0 руб. (Прямой штраф отсутствует)",
                legal_tip="Для соответствия ГОСТ Р 7.0.97-2016 измените шрифт на Times New Roman или Arial."
            ))
    return errors


def validate_order_772n(chunks: List[DocumentChunk]) -> List[ValidationErrorItem]:
    errors = []
    full_text = " ".join([c.text.lower() for c in chunks])

    required_chapters = {
        "общие требования": "Общие требования охраны труда",
        "перед началом работы": "Требования охраны труда перед началом работы",
        "во время работы": "Требования охраны труда во время работы",
        "в аварийных ситуациях": "Требования охраны труда в аварийных ситуациях",
        "по окончании работы": "Требования охраны труда по окончании работы"
    }

    for key, original_name in required_chapters.items():
        if key not in full_text:
            errors.append(ValidationErrorItem(
                category="Структура (Приказ 772н)",
                severity=SeverityLevel.CRITICAL,
                message=f"Отсутствует обязательный нормативный раздел: '{original_name}'",
                location="Structures",
                fine_equivalent="от 50 000 до 80 000 руб. (ч. 1 ст. 5.27.1 КоАП РФ)",
                legal_tip=f"Добавьте в документ главу с точным заголовком: '{original_name}'."
            ))

    if "сиз" not in full_text and "средств индивидуальной защиты" not in full_text:
        errors.append(ValidationErrorItem(
            category="Оптимизация формулировок",
            severity=SeverityLevel.RECOMMENDATION,
            message="В тексте общих требований не конкретизирован порядок выдачи и применения СИЗ.",
            location="Раздел: Общие требования",
            fine_equivalent="0 руб. (Рекомендация)",
            legal_tip="Добавьте формулировку об обязанности правильно применять выданные СИЗ."
        ))

    if "первой помощ" not in full_text and "пострадавш" not in full_text:
        errors.append(ValidationErrorItem(
            category="Качество контента",
            severity=SeverityLevel.RECOMMENDATION,
            message="Раздел аварийных ситуаций содержит слабые формулировки по оказанию первой помощи.",
            location="Раздел: В аварийных ситуациях",
            fine_equivalent="0 руб. (Рекомендация)",
            legal_tip="Рекомендуется внедрить четкий алгоритм самопомощи и взаимопомощи при обнаружении пострадавшего."
        ))

    return errors


def validate_journal_post_2464(chunks: List[DocumentChunk]) -> List[ValidationErrorItem]:
    errors = []
    full_text = " ".join([c.text.lower() for c in chunks])

    required_fields = ["фио", "дата", "подпись"]
    for field in required_fields:
        if field not in full_text:
            errors.append(ValidationErrorItem(
                category="Реквизиты документа",
                severity=SeverityLevel.CRITICAL,
                message=f"В форме фиксации инструктажа не найден обязательный столбец/поле: '{field.upper()}'",
                location="Шапка таблицы",
                fine_equivalent="от 110 000 до 130 000 руб. (ч. 3 ст. 5.27.1 КоАП РФ)",
                legal_tip=f"Без поля '{field.upper()}' ГИТ признает инструктаж недействительным."
            ))

    date_pattern = r"\b\d{2}\.\d{2}\.\d{4}\b"
    for idx, chunk in enumerate(chunks):
        match = re.search(date_pattern, chunk.text)
        if match:
            date_str = match.group(0)
            try:
                date_obj = datetime.strptime(date_str, "%d.%m.%Y")
                if datetime.now() - date_obj > timedelta(days=180):
                    errors.append(ValidationErrorItem(
                        category="Сроки действия (Пост. 2464)",
                        severity=SeverityLevel.CRITICAL,
                        message=f"Нарушена периодичность обучения! Инструктаж от {date_str} просрочен (более 6 месяцев).",
                        location=f"Строка: '{chunk.text[:30]}...'",
                        fine_equivalent="от 110 000 до 130 000 руб. ЗА КАЖДОГО сотрудника (ч. 3 ст. 5.27.1 КоАП РФ)",
                        legal_tip="Немедленно отстраните работников с просроченными датами и проведите повторный инструктаж."
                    ))
            except ValueError:
                continue

    if "номер" not in full_text and "№" not in full_text:
        errors.append(ValidationErrorItem(
            category="Structure (Постановление 2464)",
            severity=SeverityLevel.RECOMMENDATION,
            message="В таблице отсутствует сквозная нумерация записей.",
            location="Табличная часть",
            fine_equivalent="0 руб. (Рекомендация)",
            legal_tip="Добавьте первую колонку: '№ п/п' (Порядковый номер записи)."
        ))

    return errors


# === 4. ЭНДПОИНТЫ API ===

@app.post("/api/v1/verify/expert", response_model=Dict[str, Any])
async def expert_document_audit(request: ValidationRequest):
    all_violations = []

    if request.doc_type == "instruction":
        all_violations.extend(validate_gost_7_0_97(request.document_data))
        all_violations.extend(validate_order_772n(request.document_data))
    elif request.doc_type == "journal":
        all_violations.extend(validate_journal_post_2464(request.document_data))
    else:
        raise HTTPException(status_code=400, detail="Указан неподдерживаемый тип документа")

    critical_count = sum(1 for v in all_violations if v.severity == SeverityLevel.CRITICAL)

    return {
        "status": "FAILED" if critical_count > 0 else "PASSED",
        "analyzed_at": datetime.now().isoformat(),
        "total_errors": len(all_violations),
        "violations": all_violations
    }


@app.post("/api/v1/verify/export")
async def export_compliance_report(request: ValidationRequest):
    if request.doc_type not in ["instruction", "journal"]:
        raise HTTPException(status_code=400, detail="Неверный тип документа для экспорта")

    if request.doc_type == "instruction":
        violations = validate_order_772n(request.document_data) + validate_gost_7_0_97(request.document_data)
    else:
        violations = validate_journal_post_2464(request.document_data)


    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол комплаенса"
    ws.sheet_view.showGridLines = True

    headers = ["Категория", "Уровень риска", "Описание нарушения", "Локация", "Штраф по КоАП РФ",
               "Рекомендация по исправлению"]
    ws.append(headers)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_num = 2
    for v in violations:
        ws.append([
            v.category,
            v.severity.value,
            v.message,
            v.location or "Не указано",
            v.fine_equivalent,
            v.legal_tip
        ])

        severity_cell = ws.cell(row=row_num, column=2)
        if v.severity == SeverityLevel.CRITICAL:
            severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            severity_cell.font = Font(color="9C0006", bold=True)
        elif v.severity == SeverityLevel.WARNING:
            severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            severity_cell.font = Font(color="9C6500", bold=True)
        else:
            severity_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            severity_cell.font = Font(color="006100", bold=True)

        for col_num in range(1, 7):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical="top", wrap_text=True)
        row_num += 1

    # ФИКС: Безопасный перевод значений ячеек в строку для вычисления длины col
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ФИКС: Безопасное кодирование имени файла для передачи через HTTP-заголовки
    raw_filename = f"OHS_Report_{request.doc_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(raw_filename)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/api/v1/dashboard/director")
async def director_risk_dashboard():
    return {
        "company_status": "HIGH_RISK",
        "compliance_score_percent": 68,
        "git_inspection_readiness": "Не готов",
        "financial_risk_estimation": "До 1 300 000 рублей (на основе ст. 5.27.1 КоАП за системные дефекты в журнах)",
        "recommendation": "Срочно исправьте критические нарушения: добавьте недостающие разделы в инструкциях по Приказу 772н и обновите даты просроченных инструктажей в журналах."
    }